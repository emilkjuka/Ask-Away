from replit import db
import openpyxl


workbook = openpyxl.load_workbook("Questions.xlsx")
sheet = workbook.active
# build_db()

def build_db():
  for row in range(2,sheet.max_row+1):
    db[sheet.cell(row,1).value] = sheet.cell(row, 2).value
